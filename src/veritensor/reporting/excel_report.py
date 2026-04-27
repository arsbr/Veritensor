# Copyright 2026 Veritensor Security Apache 2.0
# Excel Report Generator for Auditors and Compliance Teams.
# Produces a .xlsx file with three sheets: Summary, Incidents, Suppressed.

import datetime
from pathlib import Path
from typing import List

from veritensor.core.types import ScanResult
from veritensor import __version__

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# --- Color palette ---
COLOR_HEADER_BG   = "1F3864"   # Dark navy — header rows
COLOR_HEADER_FONT = "FFFFFF"   # White text on headers
COLOR_CRITICAL    = "FF0000"   # Red
COLOR_HIGH        = "FF6600"   # Orange
COLOR_MEDIUM      = "FFD700"   # Gold
COLOR_LOW         = "4472C4"   # Blue
COLOR_PASS        = "70AD47"   # Green
COLOR_FAIL        = "FF0000"   # Red
COLOR_ALT_ROW     = "EBF3FF"   # Light blue alt row
COLOR_SUMMARY_BG  = "D6E4F0"   # Light blue for summary cells


def _severity_color(threat_str: str) -> str:
    """Returns hex color for a threat string based on its severity prefix."""
    t = threat_str.upper()
    if t.startswith("CRITICAL"):  return COLOR_CRITICAL
    if t.startswith("HIGH"):      return COLOR_HIGH
    if t.startswith("MEDIUM"):    return COLOR_MEDIUM
    if t.startswith("LOW"):       return COLOR_LOW
    return "000000"


def _worst_severity(threats: List[str]) -> str:
    """Returns the worst severity label from a list of threat strings."""
    order = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}
    worst = "LOW"
    for t in threats:
        for sev in order:
            if t.upper().startswith(sev):
                if order[sev] > order.get(worst, 0):
                    worst = sev
                break
    return worst


def _header_style(cell, bg: str = COLOR_HEADER_BG):
    cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _thin_border(cell)


def _thin_border(cell):
    side = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _data_cell(cell, value, align="left", bold=False, color=None, wrap=False):
    cell.value     = value
    cell.font      = Font(name="Arial", size=9, bold=bold,
                          color=color if color else "000000")
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    _thin_border(cell)


def _set_col_widths(ws, widths: dict):
    """widths: {column_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet 1 — Summary
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, results: List[ScanResult], scan_date: str):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "Veritensor Security Scan Report"
    title_cell.font  = Font(bold=True, name="Arial", size=14, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    total    = len(results)
    passed   = sum(1 for r in results if r.status == "PASS")
    failed   = sum(1 for r in results if r.status == "FAIL")
    critical = sum(1 for r in results
                   for t in (r.threats or []) if t.upper().startswith("CRITICAL"))
    high     = sum(1 for r in results
                   for t in (r.threats or []) if t.upper().startswith("HIGH"))
    medium   = sum(1 for r in results
                   for t in (r.threats or []) if t.upper().startswith("MEDIUM"))
    low      = sum(1 for r in results
                   for t in (r.threats or []) if t.upper().startswith("LOW"))

    rows = [
        ("Scan Date",          scan_date),
        ("Scanner Version",    f"veritensor v{__version__}"),
        ("",                   ""),
        ("Total Files Scanned", total),
        ("Passed",             passed),
        ("Failed",             failed),
        ("",                   ""),
        ("CRITICAL Threats",   critical),
        ("HIGH Threats",       high),
        ("MEDIUM Threats",     medium),
        ("LOW Threats",        low),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        a_cell = ws.cell(row=i, column=1, value=label)
        b_cell = ws.cell(row=i, column=2, value=value)

        if label == "":
            continue

        a_cell.font      = Font(bold=True, name="Arial", size=10)
        a_cell.fill      = PatternFill("solid", start_color=COLOR_SUMMARY_BG)
        a_cell.alignment = Alignment(horizontal="left", vertical="center")
        _thin_border(a_cell)

        # Color-code value cells for threat counts
        font_color = "000000"
        if label == "CRITICAL Threats" and critical > 0:  font_color = COLOR_CRITICAL
        if label == "HIGH Threats"     and high > 0:      font_color = COLOR_HIGH
        if label == "MEDIUM Threats"   and medium > 0:    font_color = COLOR_MEDIUM
        if label == "Failed"           and failed > 0:    font_color = COLOR_FAIL
        if label == "Passed"           and passed > 0:    font_color = COLOR_PASS

        b_cell.font      = Font(bold=(font_color != "000000"), name="Arial",
                                size=10, color=font_color)
        b_cell.alignment = Alignment(horizontal="center", vertical="center")
        b_cell.fill      = PatternFill("solid", start_color="F5F9FF")
        _thin_border(b_cell)

        ws.row_dimensions[i].height = 18


# ---------------------------------------------------------------------------
# Sheet 2 — Incidents (FAIL results only, one row per threat)
# ---------------------------------------------------------------------------

def _build_incidents_sheet(ws, results: List[ScanResult]):
    ws.title = "Incidents"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "File Name", "File Path", "File Hash",
        "Worst Severity", "Severity", "Threat Description",
        "License", "Identity Verified", "Repo / Source",
    ]

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _header_style(cell)
    ws.row_dimensions[1].height = 22

    row_num = 2
    for res in results:
        if res.status != "FAIL":
            continue

        threats = res.threats or []
        if not threats:
            continue

        worst = _worst_severity(threats)
        file_name = Path(res.file_path).name if res.file_path else ""
        alt_fill  = PatternFill("solid", start_color=COLOR_ALT_ROW)

        for t_idx, threat in enumerate(threats):
            # Determine per-threat severity
            sev_label = "INFO"
            for s in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
                if threat.upper().startswith(s):
                    sev_label = s
                    break

            row_data = [
                file_name,
                res.file_path or "",
                res.file_hash or "",
                worst,
                sev_label,
                threat,
                res.detected_license or "",
                "Yes" if res.identity_verified else "No",
                res.repo_id or "",
            ]

            use_alt = (row_num % 2 == 0)

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)

                if col_idx == 4:  # Worst Severity — color coded
                    _data_cell(cell, val, align="center", bold=True,
                               color=_severity_color(worst + ":"))
                elif col_idx == 5:  # Per-threat severity
                    _data_cell(cell, val, align="center",
                               color=_severity_color(threat))
                elif col_idx == 6:  # Threat description — wrap
                    _data_cell(cell, val, wrap=True)
                    ws.row_dimensions[row_num].height = 32
                elif col_idx == 8:  # Verified
                    color = COLOR_PASS if val == "Yes" else COLOR_FAIL
                    _data_cell(cell, val, align="center", color=color)
                else:
                    _data_cell(cell, val)
                    if use_alt and col_idx not in (4, 5, 8):
                        cell.fill = alt_fill

            row_num += 1

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    _set_col_widths(ws, {
        "A": 25, "B": 45, "C": 38, "D": 16,
        "E": 12, "F": 70, "G": 18, "H": 16, "I": 35,
    })


# ---------------------------------------------------------------------------
# Sheet 3 — All Files (PASS + FAIL overview)
# ---------------------------------------------------------------------------

def _build_all_files_sheet(ws, results: List[ScanResult]):
    ws.title = "All Files"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["File Name", "Status", "Worst Severity",
               "Threat Count", "License", "Hash", "Repo / Source"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _header_style(cell)
    ws.row_dimensions[1].height = 22

    for row_num, res in enumerate(results, start=2):
        threats    = res.threats or []
        status     = res.status
        worst      = _worst_severity(threats) if threats else ""
        file_name  = Path(res.file_path).name if res.file_path else ""
        alt_fill   = PatternFill("solid", start_color=COLOR_ALT_ROW)
        use_alt    = (row_num % 2 == 0)

        row_data = [
            file_name,
            status,
            worst,
            len(threats),
            res.detected_license or "",
            res.file_hash or "",
            res.repo_id or "",
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if col_idx == 2:  # Status
                color = COLOR_PASS if status == "PASS" else COLOR_FAIL
                _data_cell(cell, val, align="center", bold=True, color=color)
            elif col_idx == 3:  # Worst severity
                _data_cell(cell, val, align="center",
                           color=_severity_color(worst + ":") if worst else "000000")
            elif col_idx == 4:  # Count
                _data_cell(cell, val, align="center",
                           bold=(val > 0), color=COLOR_FAIL if val > 0 else "000000")
            else:
                _data_cell(cell, val)
                if use_alt and col_idx not in (2, 3, 4):
                    cell.fill = alt_fill

        ws.row_dimensions[row_num].height = 16

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    _set_col_widths(ws, {
        "A": 30, "B": 10, "C": 16, "D": 14,
        "E": 20, "F": 40, "G": 35,
    })


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel_report(
    results: List[ScanResult],
    output_path: str = "veritensor-report.xlsx"
) -> str:
    """
    Generates a formatted Excel report for auditors and compliance teams.
    Three sheets: Summary, Incidents (FAIL only), All Files.
    Returns the path to the saved file.
    Raises ImportError if openpyxl is not installed.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Run: pip install veritensor[excel]"
        )

    scan_date = datetime.datetime.now(datetime.timezone.utc).strftime(
        "%Y-%m-%d %H:%M:%S UTC"
    )

    wb = Workbook()

    # Remove default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_summary  = wb.create_sheet("Summary")
    ws_incidents= wb.create_sheet("Incidents")
    ws_all      = wb.create_sheet("All Files")

    _build_summary_sheet(ws_summary, results, scan_date)
    _build_incidents_sheet(ws_incidents, results)
    _build_all_files_sheet(ws_all, results)

    # Set Summary as active tab on open
    wb.active = ws_summary

    wb.save(output_path)
    return output_path
